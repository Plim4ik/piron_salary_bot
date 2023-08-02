import os
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from config import EXCEL_OUTPUT_FOLDER, EXCEL_REPORTS_FOLDER

def get_latest_file(folder):
    files = os.listdir(folder)
    excel_files = [file for file in files if file.endswith(".xlsx")]
    if not excel_files:
        return None
    sorted_files = sorted(excel_files, key=lambda x: os.path.getmtime(os.path.join(folder, x)), reverse=True)
    return os.path.join(folder, sorted_files[0])

def process_excel_file(input_file_path, output_folder):
    # Проверяем, что файл имеет расширение .xlsx
    if not input_file_path.endswith(".xlsx"):
        print(f"Ошибка: файл '{input_file_path}' не является файлом Excel (.xlsx).")
        return

    wb = openpyxl.load_workbook(input_file_path)
    sheet = wb.active

    columns_to_remove = [2, 3, 4, 6, 7, 9]
    for col_index in sorted(columns_to_remove, reverse=True):
        sheet.delete_cols(col_index)

    rows_to_delete = []
    for row in sheet.iter_rows(min_row=11, max_row=sheet.max_row):
        if row[0].value == "неотвеченный":
            rows_to_delete.append(row[0].row)

    for row_num in reversed(rows_to_delete):
        sheet.delete_rows(row_num)

    start_column = sheet.max_column + 1
    new_columns = ["Чем интересовались?", "Сайт", "Почтовик"]
    for i, column_name in enumerate(new_columns, start=0):
        col_index = start_column + i
        cell = sheet.cell(row=10, column=col_index)
        cell.value = column_name
        cell.font = Font(name="Arial", size=12, bold=True)

    date_str = datetime.today().strftime("%Y-%m-%d")
    output_file_name = f"Пирон_{date_str}_processed.xlsx"
    output_file_path = os.path.join(EXCEL_OUTPUT_FOLDER, output_file_name)
    wb.save(output_file_path)

    return output_file_path

def manage_files(folder):
    files = os.listdir(folder)
    excel_files = [file for file in files if file.endswith(".xlsx")]
    if len(excel_files) > 10:
        sorted_files = sorted(excel_files, key=lambda x: os.path.getctime(os.path.join(folder, x)))
        os.remove(os.path.join(folder, sorted_files[0]))

def handle_report_file(input_file_path):
    output_file_path = process_excel_file(input_file_path, EXCEL_REPORTS_FOLDER)
    manage_files(EXCEL_REPORTS_FOLDER)
    manage_files(EXCEL_OUTPUT_FOLDER)
    return output_file_path
